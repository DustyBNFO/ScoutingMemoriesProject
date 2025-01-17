import openpyxl
from openpyxl.styles.colors import COLOR_INDEX

# Define the path to your Excel file
excel_file_path = 'Council and Lodge directory by lodge number 06-30-2024_WIP.xlsx'  # Replace with your actual file name

# Define the RGB color values of the target colors in the form 'RRGGBB'
target_colors = ['FF9900', 'E26B0A', 'FABF8F', 'FCD5B4', 'FFC000']  # Example colors in hex format (yellow and orange)

# Load the workbook and select the active worksheet
wb = openpyxl.load_workbook(excel_file_path)
ws = wb.active

# List to store extracted information
extracted_info = []

# Function to get the color in hex
def get_hex_color(color):
    if color.type == 'rgb':
        return color.rgb[2:]  # Strip the '00' prefix
    elif color.type == 'indexed':
        if color.indexed in COLOR_INDEX:
            return COLOR_INDEX[color.indexed][2:]  # Convert to hex color and strip '00' prefix
    return None

# Function to find the main cell for merged cells
def get_merged_cell_value(sheet, cell):
    for merge in sheet.merged_cells.ranges:
        if cell.coordinate in merge:
            return sheet.cell(merge.min_row, merge.min_col)
    return cell

# Iterate through all cells in the worksheet
for row in ws.iter_rows():
    for cell in row:
        main_cell = get_merged_cell_value(ws, cell)
        fill = main_cell.fill
        if fill and fill.fgColor:
            cell_color = get_hex_color(fill.fgColor)
            if cell_color:
                print(f"Cell value: {main_cell.value}, Cell color: {cell_color}, Coordinates: {main_cell.coordinate}")  # Debugging line
                if cell_color in target_colors:
                    extracted_info.append((main_cell.value, cell_color, main_cell.coordinate))

# Print or save the extracted information
for info in extracted_info:
    print(info)

# If you want to save the extracted info to a new file
output_file_path = 'extracted_info.txt'
with open(output_file_path, 'w') as file:
    for info in extracted_info:
        file.write(f"Value: {info[0]}, Color: {info[1]}, Coordinates: {info[2]}\n")

print(f"The extracted information has been saved to '{output_file_path}'.")
