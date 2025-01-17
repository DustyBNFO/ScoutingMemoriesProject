import csv
import openpyxl

# Define the list of U.S. state abbreviations
us_state_abbreviations = set([
    'AL', 'AK', 'AZ', 'AR', 'CA', 'CO', 'CT', 'DE', 'FL', 'GA', 'HI', 'ID', 'IL', 'IN', 
    'IA', 'KS', 'KY', 'LA', 'ME', 'MD', 'MA', 'MI', 'MN', 'MS', 'MO', 'MT', 'NE', 'NV', 
    'NH', 'NJ', 'NM', 'NY', 'NC', 'ND', 'OH', 'OK', 'OR', 'PA', 'RI', 'SC', 'SD', 'TN', 
    'TX', 'UT', 'VT', 'VA', 'WA', 'WV', 'WI', 'WY'
])

# Read the CSV file and extract the required information
council_data = []

with open('virginia_only_csv.csv', mode='r', encoding='utf-8') as file:
    reader = csv.reader(file)
    
    # Read all rows and store them for later use
    rows = list(reader)
    
    # Skip the header row
    header = rows[0]
    data_rows = rows[1:]
    
    # Iterate over each row to extract the required columns
    for row in data_rows:
        name = row[2]   # 2nd column
        city = row[3]   # 3rd column
        state = row[4]  # 4th column
        start_date = row[5]  # 5th column
        end_date = row[6]    # 6th column
        
        # Append the extracted data to the list
        council_data.append((name, city, state, start_date, end_date))

# Load the Excel file
workbook = openpyxl.load_workbook('Council and Lodge directory by lodge number 06-30-2024_WIP.xlsx')
sheet = workbook.active

def find_best_matching_cell(sheet, name, city, state, start_date, end_date):
    best_match = None
    highest_match_count = 0
    
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell_value = str(cell.value).lower().replace(' ', '')
            matches_count = 0
            
            if name.lower().replace(' ', '') in cell_value:
                matches_count += 1
            if city.lower().replace(' ', '') in cell_value:
                matches_count += 1
            if state.upper().replace(' ', '') in cell_value:
                matches_count += 1
            if start_date.lower().replace(' ', '') in cell_value:
                matches_count += 1
            if end_date.lower().replace(' ', '') in cell_value:
                matches_count += 1
            
            # Update the best match if this cell has more matches
            if matches_count > highest_match_count:
                highest_match_count = matches_count
                best_match = cell
    
    return best_match

matches = []
for name, city, state, start_date, end_date in council_data:
    best_matching_cell = find_best_matching_cell(sheet, name, city, state, start_date, end_date)
    if best_matching_cell:
        matches.append((name, city, state, start_date, end_date, best_matching_cell.coordinate))

# Print the matches
for match in matches:
    print(match)

# Extract the first line and state abbreviation from the corresponding cell in column 4
extracted_lines = []
for match in matches:
    name, city, state, start_date, end_date, coordinate = match
    row = sheet[coordinate].row

    # Check for merged cells and get the value from the top-left cell of the merged range
    for merged_cell in sheet.merged_cells.ranges:
        if sheet.cell(row=row, column=4).coordinate in merged_cell:
            top_left_cell = merged_cell.min_row
            cell_in_column_4 = sheet.cell(row=top_left_cell, column=4).value
            break
    else:
        cell_in_column_4 = sheet.cell(row=row, column=4).value
    
    if cell_in_column_4:
        # Split the cell value by lines and take the first line
        cell_lines = str(cell_in_column_4).split('\n')
        first_line = cell_lines[0]
        
        # Extract the state abbreviation from the 3rd line
        state_abbreviation = ""
        if len(cell_lines) >= 3:
            third_line = cell_lines[2].strip()
            for state_abbreviation in us_state_abbreviations:
                if state_abbreviation in third_line:
                    state_abbreviation = state_abbreviation
                    break
        
        # Reformat the first line
        reformatted_council_name = first_line.replace(' ', '').replace('Council', '') + '-' + state_abbreviation
        
        extracted_lines.append(reformatted_council_name)

# Check if all reformatted council names appear at least once in the 2nd column of the original CSV
output_lines = []
for reformatted_name in extracted_lines:
    if any(reformatted_name in row[1] for row in data_rows):
        output_lines.append(reformatted_name)
    else:
        output_lines.append(f"{reformatted_name} - NOT FOUND")

# Write the output lines to a new CSV file
with open('reformatted_council_names.csv', mode='w', encoding='utf-8', newline='') as file:
    writer = csv.writer(file)
    # Write a header row
    writer.writerow(['Reformatted Council Name'])
    # Write the output lines
    for output_line in output_lines:
        writer.writerow([output_line])

print("Reformatted council names have been written to reformatted_council_names.csv with not found flags where applicable.")

